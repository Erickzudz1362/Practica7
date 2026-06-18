import os
from datetime import datetime, timezone

import requests
from fastapi import Depends, FastAPI, File, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from common.db import connect, row_to_dict, rows_to_dicts
from common.events import publish_event
from common.security import current_user


app = FastAPI(title="Inventory Service", version="1.0.0")
db = connect("inventory.db")
LOW_STOCK_THRESHOLD = 10
PRODUCT_URL = os.getenv("PRODUCT_URL", "http://127.0.0.1:8002")
COMPANY_URL = os.getenv("COMPANY_URL", "http://127.0.0.1:8001")


class InventoryInput(BaseModel):
    product_id: int
    branch_id: int
    quantity: int = Field(..., gt=0)
    cost: float = Field(0, ge=0)
    price: float = Field(..., gt=0)
    reason: str = "INGRESO"


class InventoryOutput(BaseModel):
    product_id: int
    branch_id: int
    quantity: int = Field(..., gt=0)
    reason: str = "BAJA"


class InventoryTransfer(BaseModel):
    product_id: int
    from_branch_id: int
    to_branch_id: int
    quantity: int = Field(..., gt=0)


def init_db():
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS stock (
            product_id INTEGER NOT NULL,
            branch_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 0,
            average_cost REAL NOT NULL DEFAULT 0,
            sale_price REAL NOT NULL DEFAULT 0,
            PRIMARY KEY(product_id, branch_id)
        );
        CREATE TABLE IF NOT EXISTS kardex (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            branch_id INTEGER NOT NULL,
            movement_type TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            cost REAL NOT NULL DEFAULT 0,
            price REAL NOT NULL DEFAULT 0,
            reference TEXT,
            created_at TEXT NOT NULL
        );
        """
    )
    db.commit()


init_db()


def now():
    return datetime.now(timezone.utc).isoformat()


def get_stock(product_id: int, branch_id: int):
    row = db.execute(
        "SELECT * FROM stock WHERE product_id = ? AND branch_id = ?",
        (product_id, branch_id),
    ).fetchone()
    return row_to_dict(row) or {
        "product_id": product_id,
        "branch_id": branch_id,
        "quantity": 0,
        "average_cost": 0,
        "sale_price": 0,
    }


def upsert_stock(product_id: int, branch_id: int, quantity_delta: int, cost: float = 0, price: float = 0):
    current = get_stock(product_id, branch_id)
    new_quantity = current["quantity"] + quantity_delta
    if new_quantity < 0:
        raise HTTPException(status_code=409, detail="Stock insuficiente")
    new_cost = cost if cost else current["average_cost"]
    new_price = price if price else current["sale_price"]
    db.execute(
        """
        INSERT INTO stock(product_id, branch_id, quantity, average_cost, sale_price)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(product_id, branch_id)
        DO UPDATE SET quantity = excluded.quantity,
                      average_cost = excluded.average_cost,
                      sale_price = excluded.sale_price
        """,
        (product_id, branch_id, new_quantity, new_cost, new_price),
    )
    return get_stock(product_id, branch_id) | {"quantity": new_quantity, "average_cost": new_cost, "sale_price": new_price}


def add_kardex(product_id: int, branch_id: int, movement_type: str, quantity: int, cost=0, price=0, reference=None):
    db.execute(
        """
        INSERT INTO kardex(product_id, branch_id, movement_type, quantity, cost, price, reference, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (product_id, branch_id, movement_type, quantity, cost, price, reference, now()),
    )


def auth_headers(request: Request):
    return {"Authorization": request.headers.get("Authorization", "")}


def get_json_or_none(url: str, request: Request):
    try:
        response = requests.get(url, headers=auth_headers(request), timeout=5)
        if response.status_code >= 400:
            return None
        return response.json()
    except requests.RequestException:
        return None


@app.get("/health")
def health():
    return {"status": "ok", "service": "inventory-service"}


@app.post("/inventory/loadExcel", dependencies=[Depends(current_user)])
async def load_excel(file: UploadFile = File(...)):
    workbook = load_workbook(file.file)
    sheet = workbook.active
    headers = [str(cell.value).strip().lower() for cell in sheet[1]]
    required = {"producto", "sucursal", "cantidad", "costo", "precio"}
    if not required.issubset(set(headers)):
        raise HTTPException(status_code=400, detail="El Excel debe incluir Producto, Sucursal, Cantidad, Costo y Precio")
    index = {name: headers.index(name) for name in headers}
    loaded = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[index["producto"]] is None:
            continue
        item = InventoryInput(
            product_id=int(row[index["producto"]]),
            branch_id=int(row[index["sucursal"]]),
            quantity=int(row[index["cantidad"]]),
            cost=float(row[index["costo"]]),
            price=float(row[index["precio"]]),
            reason="IMPORTACION_EXCEL",
        )
        stock = input_inventory(item, publish=False)
        loaded.append(stock)
    db.commit()
    publish_event("InventoryLoaded", {"items": loaded})
    return {"loaded": len(loaded), "items": loaded}


@app.post("/inventory/input", dependencies=[Depends(current_user)])
def input_inventory(request: InventoryInput, publish: bool = True):
    stock = upsert_stock(request.product_id, request.branch_id, request.quantity, request.cost, request.price)
    add_kardex(request.product_id, request.branch_id, "INPUT", request.quantity, request.cost, request.price, request.reason)
    db.commit()
    if publish:
        publish_event("InventoryUpdated", stock)
    return stock


@app.post("/inventory/output", dependencies=[Depends(current_user)])
def output_inventory(request: InventoryOutput):
    stock = upsert_stock(request.product_id, request.branch_id, -request.quantity)
    add_kardex(request.product_id, request.branch_id, "OUTPUT", -request.quantity, reference=request.reason)
    db.commit()
    publish_event("InventoryUpdated", stock)
    if stock["quantity"] <= LOW_STOCK_THRESHOLD:
        publish_event("StockLow", stock)
    return stock


@app.post("/inventory/reserve", dependencies=[Depends(current_user)])
def reserve_for_sale(request: InventoryOutput):
    request.reason = f"VENTA:{request.reason}"
    return output_inventory(request)


@app.post("/inventory/transfer", dependencies=[Depends(current_user)])
def transfer_inventory(request: InventoryTransfer):
    origin = upsert_stock(request.product_id, request.from_branch_id, -request.quantity)
    destination_price = origin["sale_price"]
    destination = upsert_stock(request.product_id, request.to_branch_id, request.quantity, origin["average_cost"], destination_price)
    add_kardex(request.product_id, request.from_branch_id, "TRANSFER_OUT", -request.quantity, reference=str(request.to_branch_id))
    add_kardex(request.product_id, request.to_branch_id, "TRANSFER_IN", request.quantity, origin["average_cost"], destination_price, str(request.from_branch_id))
    db.commit()
    payload = {"origin": origin, "destination": destination, "quantity": request.quantity}
    publish_event("TransferCompleted", payload)
    return payload


@app.get("/inventory/balance", dependencies=[Depends(current_user)])
def balance(product_id: int | None = None, branch_id: int | None = None):
    sql = "SELECT * FROM stock WHERE 1=1"
    params = []
    if product_id:
        sql += " AND product_id = ?"
        params.append(product_id)
    if branch_id:
        sql += " AND branch_id = ?"
        params.append(branch_id)
    sql += " ORDER BY product_id, branch_id"
    return rows_to_dicts(db.execute(sql, params).fetchall())


@app.get("/inventory/report/consolidated/{product_id}", dependencies=[Depends(current_user)])
def consolidated_product(product_id: int, request: Request):
    rows = rows_to_dicts(db.execute("SELECT * FROM stock WHERE product_id = ? ORDER BY branch_id", (product_id,)).fetchall())
    product = get_json_or_none(f"{PRODUCT_URL}/products/{product_id}", request) or {"id": product_id, "name": None}
    branches = get_json_or_none(f"{COMPANY_URL}/branches", request) or []
    branch_map = {branch["id"]: branch for branch in branches}
    enriched = []
    for row in rows:
        branch = branch_map.get(row["branch_id"], {})
        enriched.append(
            row
            | {
                "product_name": product.get("name"),
                "branch_name": branch.get("name"),
                "company_name": branch.get("company_name"),
                "city_name": branch.get("city_name"),
            }
        )
    return {
        "product_id": product_id,
        "product_name": product.get("name"),
        "total_quantity": sum(row["quantity"] for row in rows),
        "branches": enriched,
    }


@app.get("/inventory/kardex/{product_id}", dependencies=[Depends(current_user)])
def kardex(product_id: int):
    rows = db.execute("SELECT * FROM kardex WHERE product_id = ? ORDER BY id", (product_id,)).fetchall()
    return rows_to_dicts(rows)


@app.get("/inventory/{product_id}", dependencies=[Depends(current_user)])
def product_stock(product_id: int):
    return rows_to_dicts(db.execute("SELECT * FROM stock WHERE product_id = ? ORDER BY branch_id", (product_id,)).fetchall())
